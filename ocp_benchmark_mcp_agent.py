#!/usr/bin/env python3
"""OpenShift Benchmark MCP AI Agent using StateGraph for performance analysis and reporting."""
import asyncio
import json
import logging
import os
import sys
from datetime import datetime, timezone
from pathlib import Path
from typing import Dict, Any, List, Optional, TypedDict

# Add project root to path
project_root = Path(__file__).parent
sys.path.insert(0, str(project_root))

import httpx
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from fpdf import FPDF
from langchain_openai import ChatOpenAI
from langchain_core.messages import HumanMessage, AIMessage, SystemMessage
from langgraph.graph import StateGraph, END
from langgraph.checkpoint.memory import MemorySaver
from analysis.ocp_benchmark_performance_analysis import PerformanceAnalyzer, analyze_comprehensive_performance
from elt.ocp_benchmark_elt import BenchmarkDataProcessor

from dotenv import load_dotenv
#MCP imports
from mcp import ClientSession
from mcp.client.streamable_http import streamablehttp_client


# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# Set timezone to UTC
os.environ['TZ'] = 'UTC'
load_dotenv()

# Load environment variables
base_url = os.getenv("BASE_URL")    
api_key = os.getenv("GEMINI_API_KEY")
analysis_results = {}
class AgentState(TypedDict):
    """State for the LangGraph agent."""
    messages: List[Dict[str, Any]]
    current_task: str
    collected_data: Dict[str, Any]
    analysis_results: Dict[str, Any]
    report_generated: bool
    error: Optional[str]


class OCPBenchmarkAgent:
    """AI Agent for OpenShift performance analysis and reporting."""
    
    def __init__(self, openai_api_key: str, mcp_server_url: str = "http://localhost:8000"):
        self.llm = ChatOpenAI(
            model="gemini-1.5-flash",
            base_url=base_url,
            api_key=api_key,
            temperature=0.1,
            streaming=True         
        )
        
        self.mcp_server_url = mcp_server_url
        self.analyzer = PerformanceAnalyzer()
        self.processor = BenchmarkDataProcessor()
        self.checkpointer = MemorySaver()
        self.graph = self._create_graph()
    
    def _create_graph(self) -> StateGraph:
        """Create the LangGraph state graph."""
        graph = StateGraph(AgentState)
        
        # Add nodes
        graph.add_node("data_collection", self._collect_data)
        graph.add_node("data_analysis", self._analyze_data)
        graph.add_node("report_generation", self._generate_report)
        graph.add_node("error_handling", self._handle_error)
        
        # Set entry point
        graph.set_entry_point("data_collection")
        
        # Add edges
        graph.add_conditional_edges(
            "data_collection",
            self._should_continue_to_analysis,
            {
                "analysis": "data_analysis",
                "error": "error_handling"
            }
        )
        
        graph.add_conditional_edges(
            "data_analysis",
            self._should_continue_to_report,
            {
                "report": "report_generation",
                "error": "error_handling"
            }
        )
        
        graph.add_edge("report_generation", END)
        graph.add_edge("error_handling", END)
        
        return graph.compile(checkpointer=self.checkpointer)
    
    async def _collect_data(self, state: AgentState) -> AgentState:
        """Collect performance data from MCP server."""
        logger.info("Starting data collection...")
        
        try:
            collected_data = {}
            
            # Collect cluster information
            cluster_info = await self._call_mcp_tool("get_cluster_information")
            collected_data["cluster_info"] = cluster_info
            
            # Collect node information
            node_info = await self._call_mcp_tool("get_node_information")
            collected_data["node_info"] = node_info
            
            # Collect nodes usage metrics
            nodes_usage = await self._call_mcp_tool("get_nodes_usage_metrics", {
                "duration_hours": 2.0,  # Collect 2 hours of data
                "step": "1m"
            })
            collected_data["nodes_usage"] = nodes_usage

            # Collect pods usage metrics
            pods_usage = await self._call_mcp_tool("get_pods_usage_metrics", {
                "pod_regex": "etcd.*|ovnkube-.*|kube-apiserver.*",
                "duration_hours": 1.0,
                "step": "5m"
            })
            collected_data["pods_usage"] = pods_usage
            
            # Collect disk metrics
            disk_metrics = await self._call_mcp_tool("get_disk_io_metrics", {
                "duration_hours": 1.0,
                "step": "1m"
            })
            collected_data["disk_metrics"] = disk_metrics
            
            # Collect network metrics
            network_metrics = await self._call_mcp_tool("get_network_performance_metrics", {
                "duration_hours": 1.0,
                "step": "1m"
            })
            collected_data["network_metrics"] = network_metrics
            
            # Collect API latency metrics
            api_request_latency_metrics = await self._call_mcp_tool("get_api_request_latency_metrics", {
                "duration_hours": 1.0,
                "step": "1m"
            })
            collected_data["api_metrics"] = api_request_latency_metrics


            # Collect API latency metrics
            api_request_rate_metrics = await self._call_mcp_tool("get_api_request_latency_metrics", {
                "duration_hours": 1.0,
                "step": "1m"
            })
            collected_data["api_metrics"] = api_request_rate_metrics

            # Collect API latency metrics
            etcd_latency_metrics = await self._call_mcp_tool("get_etcd_latency_metrics", {
                "duration_hours": 1.0,
                "step": "1m"
            })
            collected_data["api_metrics"] = etcd_latency_metrics            

            state["collected_data"] = collected_data
            state["messages"].append({
                "type": "system",
                "content": f"Successfully collected {len(collected_data)} datasets"
            })
            
            logger.info(f"Data collection completed: {len(collected_data)} datasets")
            
        except Exception as e:
            logger.error(f"Data collection failed: {e}")
            state["error"] = f"Data collection failed: {str(e)}"
        
        return state
    
    async def _analyze_data(self, state: AgentState) -> AgentState:
        """Analyze the collected performance data."""
        logger.info("Starting data analysis...")
        
        try:
            collected_data = state.get("collected_data", {})
            if not collected_data:
                raise ValueError("No data to analyze")
            
            # Perform comprehensive analysis on each dataset
            global analysis_results
            
            # Analyze nodes usage
            if "nodes_usage" in collected_data:
                nodes_data = self._parse_json_response(collected_data["nodes_usage"])
                nodes_analysis = analyze_comprehensive_performance(nodes_data)

                analysis_results["nodes_analysis"] = self._parse_json_response(nodes_analysis)
            
            # Analyze pods usage
            if "pods_usage" in collected_data:
                pods_data = self._parse_json_response(collected_data["pods_usage"])
                pods_analysis = self.processor.analyze_performance_against_baselines(pods_data)
                analysis_results["pods_analysis"] = pods_analysis
            
            # Analyze disk metrics
            if "disk_metrics" in collected_data:
                disk_data = self._parse_json_response(collected_data["disk_metrics"])
                disk_analysis = analyze_comprehensive_performance(disk_data)
                analysis_results["disk_analysis"] = self._parse_json_response(disk_analysis)
            
            # Analyze network metrics
            if "network_metrics" in collected_data:
                network_data = self._parse_json_response(collected_data["network_metrics"])
                network_analysis = analyze_comprehensive_performance(network_data)
                analysis_results["network_analysis"] = self._parse_json_response(network_analysis)
            
            # Analyze API metrics
            if "api_metrics" in collected_data:
                api_data = self._parse_json_response(collected_data["api_metrics"])
                api_analysis = analyze_comprehensive_performance(api_data)
                analysis_results["api_analysis"] = self._parse_json_response(api_analysis)
            
            # Generate overall cluster health summary
            overall_summary = self._generate_overall_summary(analysis_results)
            analysis_results["overall_summary"] = overall_summary
            
            state["analysis_results"] = analysis_results
            state["messages"].append({
                "type": "system",
                "content": f"Analysis completed for {len(analysis_results)} components"
            })
            
            logger.info("Data analysis completed")
            
        except Exception as e:
            logger.error(f"Data analysis failed: {e}")
            state["error"] = f"Data analysis failed: {str(e)}"
        
        return state
    
    async def _generate_report(self, state: AgentState) -> AgentState:
        """Generate Excel and PDF reports."""
        logger.info("Starting report generation...")
        
        try:
            analysis_results = state.get("analysis_results", {})
            collected_data = state.get("collected_data", {})
            
            if not analysis_results:
                raise ValueError("No analysis results to report")
            
            # Generate timestamp for filenames
            timestamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
            
            # Generate Excel report
            excel_path = await self._generate_excel_report(
                analysis_results, 
                collected_data, 
                f"exports/ocp_benchmark_report_{timestamp}.xlsx"
            )
            
            # Generate PDF report
            pdf_path = await self._generate_pdf_report(
                analysis_results,
                f"exports/ocp_benchmark_report_{timestamp}.pdf"
            )
            
            state["report_generated"] = True
            state["messages"].append({
                "type": "system",
                "content": f"Reports generated successfully:\n- Excel: {excel_path}\n- PDF: {pdf_path}"
            })
            
            logger.info(f"Report generation completed: {excel_path}, {pdf_path}")
            
        except Exception as e:
            logger.error(f"Report generation failed: {e}")
            state["error"] = f"Report generation failed: {str(e)}"
        
        return state
    
    def _should_continue_to_analysis(self, state: AgentState) -> str:
        """Decide whether to continue to analysis or handle error."""
        return "error" if state.get("error") else "analysis"
    
    def _should_continue_to_report(self, state: AgentState) -> str:
        """Decide whether to continue to report generation or handle error."""
        return "error" if state.get("error") else "report"
    
    async def _handle_error(self, state: AgentState) -> AgentState:
        """Handle errors in the workflow."""
        error_msg = state.get("error", "Unknown error")
        logger.error(f"Agent error: {error_msg}")
        
        state["messages"].append({
            "type": "error",
            "content": f"Workflow failed: {error_msg}"
        })
        
        return state
    
    async def _call_mcp_tool(self, tool_name: str, params: Optional[Dict[str, Any]] = None) -> Dict[str, Any]:

        """Call a tool on the MCP tool via HTTP.

        Args:
            tool_name: Name of the tool to call
            params: Parameters for the tool

        Returns:
            Tool response data
        """    
        try:
            url = f"{self.mcp_server_url}/mcp"

            # Connect to the server using Streamable HTTP
            async with streamablehttp_client(
                url,
                # headers={"accept": "application/json"}
                ) as (
                    read_stream,
                    write_stream,
                    get_session_id,
            ):
                async with ClientSession(read_stream, write_stream) as session:
                    # Initialize the connection
                    await session.initialize()
 
                    # Get session id once connection established
                    session_id = get_session_id() 
                    print("Session ID: in call_tool", session_id)
            
                    print(f"Calling tool {tool_name} with params {params}", type(params))

                    # Call tool with parameters under the required 'params' key to satisfy server schema
                    result = await session.call_tool(tool_name, {"params": params or {}})
                    # print("#*"*50)
                    # print("result in call_tool of mcp client:",result)
                    # print("#*"*50)
                    # print(f"{tool_name} = {result.content[0].text}")
                    json_data = json.loads(result.content[0].text)
                    return json_data

        except Exception as e:
            print(f"Call MCP tool failed: {str(e)}")
            import traceback
            traceback.print_exc()
            raise
    
    def _parse_json_response(self, response: Dict[str, Any]) -> Dict[str, Any]:
        """Parse JSON response from MCP tools."""
        if isinstance(response, dict) and "result" in response:
            result = response["result"]
            if isinstance(result, str):
                try:
                    return json.loads(result)
                except json.JSONDecodeError:
                    return {"raw_response": result}
            return result
        return response
    
    def _generate_overall_summary(self, analysis_results: Dict[str, Any]) -> Dict[str, Any]:
        """Generate overall cluster health summary."""
        summary = {
            "timestamp": datetime.now(timezone.utc).isoformat(),
            "overall_health_score": 0,
            "health_status": "unknown",
            "critical_issues": [],
            "warning_issues": [],
            "recommendations": [],
            "component_scores": {}
        }
 
        # Collect health scores from different components
        health_scores = []
        all_recommendations = set()
        
        for component, analysis in analysis_results.items():
            parsed = analysis
            # Parse JSON string analyses
            if isinstance(parsed, str):
                try:
                    parsed = json.loads(parsed)
                except Exception:
                    parsed = {}
            # Ensure dict for extraction
            if not isinstance(parsed, dict):
                continue

            score: float = None  # type: ignore[assignment]
            issues = []
            recs = []

            if "health_analysis" in parsed and isinstance(parsed["health_analysis"], dict):
                health_data = parsed["health_analysis"]
                score = health_data.get("overall_health_score")
                issues = health_data.get("issues", []) or []
                recs = health_data.get("recommendations", []) or []
            elif "summary" in parsed and isinstance(parsed["summary"], dict):
                # Fallback when only a summary with health_score is available
                score = parsed["summary"].get("health_score")

            if isinstance(score, (int, float)):
                health_scores.append(float(score))
                summary["component_scores"][component] = float(score)

            # Collect issues
            for issue in issues:
                try:
                    text = str(issue)
                except Exception:
                    text = ""
                if not text:
                    continue
                if "critical" in text.lower():
                    summary["critical_issues"].append(f"{component}: {text}")
                else:
                    summary["warning_issues"].append(f"{component}: {text}")

            # Collect recommendations
            for rec in recs:
                all_recommendations.add(rec)
        
        # Calculate overall health score
        if health_scores:
            summary["overall_health_score"] = round(sum(health_scores) / len(health_scores), 1)
            
            # Determine health status
            avg_score = summary["overall_health_score"]
            if avg_score >= 90:
                summary["health_status"] = "excellent"
            elif avg_score >= 75:
                summary["health_status"] = "good"
            elif avg_score >= 60:
                summary["health_status"] = "fair"
            elif avg_score >= 40:
                summary["health_status"] = "poor"
            else:
                summary["health_status"] = "critical"
        
        summary["recommendations"] = list(all_recommendations)
        
        return summary
    
    async def _generate_excel_report(self, analysis_results: Dict[str, Any], 
                                   collected_data: Dict[str, Any], 
                                   filename: str) -> str:
        """Generate Excel report."""
        # Create exports directory if it doesn't exist
        Path("exports").mkdir(exist_ok=True)
        
        wb = Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Create summary sheet
        summary_sheet = wb.create_sheet("Executive Summary")
        self._create_summary_sheet(summary_sheet, analysis_results)
        
        # Create detailed sheets for each component
        # Always create Nodes sheet; fall back to raw nodes_usage if analysis is missing
        nodes_sheet = wb.create_sheet("Nodes Analysis")
        self._create_nodes_sheet(
            nodes_sheet,
            analysis_results.get("nodes_analysis", {}),
            collected_data.get("nodes_usage")
        )
        
        if "pods_analysis" in analysis_results:
            pods_sheet = wb.create_sheet("Pods Analysis")
            self._create_pods_sheet(pods_sheet, analysis_results["pods_analysis"], collected_data.get("pods_usage"))

        if "network_analysis" in analysis_results:
            network_sheet = wb.create_sheet("Network Analysis")
            self._create_network_sheet(network_sheet, analysis_results["network_analysis"], collected_data.get("network_metrics"))
        
        # Create recommendations sheet
        recs_sheet = wb.create_sheet("Recommendations")
        self._create_recommendations_sheet(recs_sheet, analysis_results)
        
        # Save workbook
        wb.save(filename)
        logger.info(f"Excel report saved: {filename}")
        
        return filename
    
    def _create_summary_sheet(self, sheet, analysis_results: Dict[str, Any]):
        """Create executive summary sheet in Excel."""
        # Title
        sheet['A1'] = "OpenShift Cluster Performance Report"
        sheet['A1'].font = Font(size=16, bold=True)
        
        # Report metadata
        sheet['A3'] = "Generated:"
        sheet['B3'] = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC")
        
        overall_summary = analysis_results.get("overall_summary", {})
        
        # Overall health
        sheet['A5'] = "Overall Health Score:"
        sheet['B5'] = overall_summary.get("overall_health_score", "Unknown")
        
        sheet['A6'] = "Health Status:"
        sheet['B6'] = overall_summary.get("health_status", "Unknown").title()
        
        # Component scores
        sheet['A8'] = "Component Health Scores:"
        row = 9
        for component, score in overall_summary.get("component_scores", {}).items():
            sheet[f'A{row}'] = component.replace("_", " ").title()
            sheet[f'B{row}'] = score
            row += 1
        
        # Issues summary
        sheet[f'A{row + 1}'] = "Critical Issues:"
        row += 2
        for issue in overall_summary.get("critical_issues", []):
            sheet[f'A{row}'] = issue
            sheet[f'A{row}'].fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
            row += 1
        
        # Apply formatting
        for col in ['A', 'B']:
            sheet.column_dimensions[col].width = 30
    
    def _create_nodes_sheet(self, sheet, nodes_analysis: Dict[str, Any], raw_data: Optional[Dict[str, Any]]):
        """Create nodes analysis sheet."""
        sheet['A1'] = "Nodes Performance Analysis"
        sheet['A1'].font = Font(size=14, bold=True)
        
        # Write analysis summary if available
        current_row = 3
        if isinstance(nodes_analysis, dict) and nodes_analysis:
            health = nodes_analysis.get('health_analysis', {}) or {}
            bench = nodes_analysis.get('benchmark_comparison', {}) or {}
            summary = nodes_analysis.get('summary', {}) or {}

            health_score = health.get('overall_health_score', summary.get('health_score', ''))
            health_status = str(health.get('health_status', summary.get('overall_health', ''))).title()
            benchmark_rating = bench.get('overall_rating', summary.get('benchmark_rating', ''))

            # Analysis timestamp
            analysis_ts = nodes_analysis.get('analysis_timestamp', '')
            if analysis_ts:
                sheet[f'A{current_row}'] = 'Analysis Timestamp:'
                sheet[f'B{current_row}'] = analysis_ts
                current_row += 1

            if health_score != '':
                sheet[f'A{current_row}'] = 'Overall Health Score:'
                sheet[f'B{current_row}'] = health_score
                current_row += 1
            if health_status:
                sheet[f'A{current_row}'] = 'Health Status:'
                sheet[f'B{current_row}'] = health_status
                current_row += 1
            if benchmark_rating:
                sheet[f'A{current_row}'] = 'Benchmark Rating:'
                sheet[f'B{current_row}'] = benchmark_rating
                current_row += 2

            # Health analysis details
            if health:
                sheet[f'A{current_row}'] = 'Health Analysis'
                sheet[f'A{current_row}'].font = Font(bold=True)
                current_row += 1

                # Component scores table
                comp_scores = health.get('component_scores', {}) or {}
                if comp_scores:
                    sheet[f'A{current_row}'] = 'Component'
                    sheet[f'B{current_row}'] = 'Score'
                    sheet[f'A{current_row}'].font = Font(bold=True)
                    sheet[f'B{current_row}'].font = Font(bold=True)
                    current_row += 1
                    for comp, score in comp_scores.items():
                        sheet[f'A{current_row}'] = str(comp).replace('_', ' ').title()
                        sheet[f'B{current_row}'] = score
                        current_row += 1
                    current_row += 1

                # Issues list
                issues = health.get('issues', []) or []
                sheet[f'A{current_row}'] = 'Issues:'
                sheet[f'A{current_row}'].font = Font(bold=True)
                current_row += 1
                if issues:
                    for issue in issues[:20]:
                        sheet[f'A{current_row}'] = str(issue)
                        current_row += 1
                else:
                    sheet[f'A{current_row}'] = 'None'
                    current_row += 1
                current_row += 1

            # Capacity forecast
            capacity = nodes_analysis.get('capacity_forecast', {}) or {}
            if capacity:
                sheet[f'A{current_row}'] = 'Capacity Forecast'
                sheet[f'A{current_row}'].font = Font(bold=True)
                current_row += 1
                if 'forecast_period_days' in capacity:
                    sheet[f'A{current_row}'] = 'Forecast Period (days):'
                    sheet[f'B{current_row}'] = capacity.get('forecast_period_days')
                    current_row += 1
                if 'confidence' in capacity:
                    sheet[f'A{current_row}'] = 'Confidence:'
                    sheet[f'B{current_row}'] = capacity.get('confidence')
                    current_row += 2

            # Benchmark comparison
            if bench:
                sheet[f'A{current_row}'] = 'Benchmark Comparison'
                sheet[f'A{current_row}'].font = Font(bold=True)
                current_row += 1
                if 'overall_rating' in bench:
                    sheet[f'A{current_row}'] = 'Overall Rating:'
                    sheet[f'B{current_row}'] = bench.get('overall_rating')
                    current_row += 1
                if 'benchmark_date' in bench:
                    sheet[f'A{current_row}'] = 'Benchmark Date:'
                    sheet[f'B{current_row}'] = bench.get('benchmark_date')
                    current_row += 2

            # Summary block
            if summary:
                sheet[f'A{current_row}'] = 'Summary'
                sheet[f'A{current_row}'].font = Font(bold=True)
                current_row += 1
                if 'overall_health' in summary:
                    sheet[f'A{current_row}'] = 'Overall Health:'
                    sheet[f'B{current_row}'] = str(summary.get('overall_health', '')).title()
                    current_row += 1
                if 'health_score' in summary:
                    sheet[f'A{current_row}'] = 'Health Score:'
                    sheet[f'B{current_row}'] = summary.get('health_score')
                    current_row += 1
                if 'total_recommendations' in summary:
                    sheet[f'A{current_row}'] = 'Total Recommendations:'
                    sheet[f'B{current_row}'] = summary.get('total_recommendations')
                    current_row += 1
                if 'high_priority_recommendations' in summary:
                    sheet[f'A{current_row}'] = 'High Priority Recommendations:'
                    sheet[f'B{current_row}'] = summary.get('high_priority_recommendations')
                    current_row += 2

            # Key recommendations (brief)
            recs = nodes_analysis.get('optimization_recommendations', []) or []
            if recs:
                sheet[f'A{current_row}'] = 'Key Recommendations:'
                sheet[f'A{current_row}'].font = Font(bold=True)
                current_row += 1
                sheet[f'A{current_row}'] = 'Priority'
                sheet[f'B{current_row}'] = 'Recommendation'
                sheet[f'A{current_row}'].font = Font(bold=True)
                sheet[f'B{current_row}'].font = Font(bold=True)
                current_row += 1

                for rec in recs[:10]:
                    sheet[f'A{current_row}'] = rec.get('priority', '')
                    sheet[f'B{current_row}'] = rec.get('recommendation', '')
                    current_row += 1
                current_row += 1

        # Parse raw data to extract node metrics
        if raw_data and isinstance(raw_data, dict):
            nodes_data = self._parse_json_response(raw_data) or {}

            # Write cluster statistics and baseline comparison if present
            if isinstance(nodes_data, dict):
                cluster_stats = nodes_data.get('cluster_statistics') or {}
                baseline_comp = nodes_data.get('baseline_comparison') or {}

                if cluster_stats:
                    sheet[f'A{current_row}'] = 'Cluster Statistics'
                    sheet[f'A{current_row}'].font = Font(bold=True)
                    current_row += 1

                    cpu_stats = cluster_stats.get('cpu_usage') or {}
                    mem_stats = cluster_stats.get('memory_usage') or {}

                    sheet[f'A{current_row}'] = 'CPU Min (%)'
                    sheet[f'B{current_row}'] = cpu_stats.get('min')
                    current_row += 1
                    sheet[f'A{current_row}'] = 'CPU Max (%)'
                    sheet[f'B{current_row}'] = cpu_stats.get('max')
                    current_row += 1
                    sheet[f'A{current_row}'] = 'CPU Mean (%)'
                    sheet[f'B{current_row}'] = cpu_stats.get('mean')
                    current_row += 1
                    sheet[f'A{current_row}'] = 'CPU Samples'
                    sheet[f'B{current_row}'] = cpu_stats.get('count')
                    current_row += 1

                    sheet[f'A{current_row}'] = 'Memory Min (%)'
                    sheet[f'B{current_row}'] = mem_stats.get('min')
                    current_row += 1
                    sheet[f'A{current_row}'] = 'Memory Max (%)'
                    sheet[f'B{current_row}'] = mem_stats.get('max')
                    current_row += 1
                    sheet[f'A{current_row}'] = 'Memory Mean (%)'
                    sheet[f'B{current_row}'] = mem_stats.get('mean')
                    current_row += 1
                    sheet[f'A{current_row}'] = 'Memory Samples'
                    sheet[f'B{current_row}'] = mem_stats.get('count')
                    current_row += 1

                    sheet[f'A{current_row}'] = 'Total Nodes'
                    sheet[f'B{current_row}'] = cluster_stats.get('total_nodes')
                    current_row += 1
                    sheet[f'A{current_row}'] = 'Query Duration (hours)'
                    sheet[f'B{current_row}'] = cluster_stats.get('query_duration_hours')
                    current_row += 2

                if baseline_comp:
                    sheet[f'A{current_row}'] = 'Baseline Comparison'
                    sheet[f'A{current_row}'].font = Font(bold=True)
                    current_row += 1

                    cpu_base = baseline_comp.get('cpu') or {}
                    sheet[f'A{current_row}'] = 'CPU Current Mean (%)'
                    sheet[f'B{current_row}'] = cpu_base.get('current_mean')
                    current_row += 1
                    sheet[f'A{current_row}'] = 'CPU Current Max (%)'
                    sheet[f'B{current_row}'] = cpu_base.get('current_max')
                    current_row += 1
                    sheet[f'A{current_row}'] = 'CPU Baseline Mean (%)'
                    sheet[f'B{current_row}'] = cpu_base.get('baseline_mean')
                    current_row += 1
                    sheet[f'A{current_row}'] = 'CPU Baseline Max (%)'
                    sheet[f'B{current_row}'] = cpu_base.get('baseline_max')
                    current_row += 1
                    sheet[f'A{current_row}'] = 'CPU Status'
                    sheet[f'B{current_row}'] = cpu_base.get('status')
                    current_row += 1
                    sheet[f'A{current_row}'] = 'CPU Message'
                    sheet[f'B{current_row}'] = cpu_base.get('message')
                    current_row += 1

                    mem_base = baseline_comp.get('memory') or {}
                    sheet[f'A{current_row}'] = 'Memory Current Mean (%)'
                    sheet[f'B{current_row}'] = mem_base.get('current_mean')
                    current_row += 1
                    sheet[f'A{current_row}'] = 'Memory Current Max (%)'
                    sheet[f'B{current_row}'] = mem_base.get('current_max')
                    current_row += 1
                    sheet[f'A{current_row}'] = 'Memory Baseline Mean (%)'
                    sheet[f'B{current_row}'] = mem_base.get('baseline_mean')
                    current_row += 1
                    sheet[f'A{current_row}'] = 'Memory Baseline Max (%)'
                    sheet[f'B{current_row}'] = mem_base.get('baseline_max')
                    current_row += 1
                    sheet[f'A{current_row}'] = 'Memory Status'
                    sheet[f'B{current_row}'] = mem_base.get('status')
                    current_row += 1
                    sheet[f'A{current_row}'] = 'Memory Message'
                    sheet[f'B{current_row}'] = mem_base.get('message')
                    current_row += 2


            # Normalize different possible schemas to the expected {'nodes': {...}} format
            normalized: Dict[str, Any] = {}
            if isinstance(nodes_data, dict):
                if isinstance(nodes_data.get("nodes"), dict):
                    normalized = nodes_data  # already in expected format
                elif isinstance(nodes_data.get("node_statistics"), dict):
                    normalized = {
                        "nodes": nodes_data.get("node_statistics", {})
                    }
                    # Optionally propagate timestamp and collection period
                    if "timestamp" in nodes_data:
                        normalized["timestamp"] = nodes_data.get("timestamp")
                    query_hours = (
                        (nodes_data.get("cluster_statistics") or {}).get("query_duration_hours")
                        if isinstance(nodes_data.get("cluster_statistics"), dict) else None
                    )
                    if query_hours is not None:
                        normalized["collection_period"] = {"duration_hours": query_hours}

            if normalized.get("nodes"):
                df = self.processor.nodes_usage_to_dataframe(normalized)
                if df is not None and not df.empty:
                    # Section header for details
                    sheet[f'A{current_row}'] = 'Nodes Usage Details'
                    sheet[f'A{current_row}'].font = Font(bold=True)
                    current_row += 1
                    
                    # Add headers
                    headers = list(df.columns)
                    for col, header in enumerate(headers, 1):
                        sheet.cell(row=current_row, column=col, value=header)
                        sheet.cell(row=current_row, column=col).font = Font(bold=True)
                    
                    # Add data
                    for idx, (_, data_row) in enumerate(df.iterrows(), current_row + 1):
                        for col, value in enumerate(data_row, 1):
                            sheet.cell(row=idx, column=col, value=value)
                    return
        
        # # Fallback placeholder when no nodes data is available
        if not raw_data:
           sheet['A3'] = "No nodes usage data available"
    
    def _create_pods_sheet(self, sheet, pods_analysis: Dict[str, Any], raw_data: Optional[Dict[str, Any]]):
        """Create pods analysis sheet."""
        sheet['A1'] = "Pods Performance Analysis"
        sheet['A1'].font = Font(size=14, bold=True)

        # Write pods_analysis summary at top
        current_row = 3
        if isinstance(pods_analysis, dict) and pods_analysis:
            ts = pods_analysis.get('timestamp')
            if ts:
                sheet[f'A{current_row}'] = 'Analysis Timestamp:'
                sheet[f'B{current_row}'] = ts
                current_row += 1

            overall_status = pods_analysis.get('overall_status')
            if overall_status:
                sheet[f'A{current_row}'] = 'Overall Status:'
                sheet[f'B{current_row}'] = str(overall_status).title()
                current_row += 1

            # Findings
            findings = pods_analysis.get('findings') or []
            sheet[f'A{current_row}'] = 'Findings:'
            sheet[f'A{current_row}'].font = Font(bold=True)
            current_row += 1
            if findings:
                for item in findings[:20]:
                    sheet[f'A{current_row}'] = str(item)
                    current_row += 1
            else:
                sheet[f'A{current_row}'] = 'None'
                current_row += 1
            current_row += 1

            # Recommendations
            recs = pods_analysis.get('recommendations') or []
            if recs:
                sheet[f'A{current_row}'] = 'Recommendations:'
                sheet[f'A{current_row}'].font = Font(bold=True)
                current_row += 1
                for rec in recs[:15]:
                    sheet[f'A{current_row}'] = str(rec)
                    current_row += 1
                current_row += 1

            # Baseline comparisons (kept as a placeholder section if provided)
            base = pods_analysis.get('baseline_comparisons') or {}
            if isinstance(base, dict) and base:
                sheet[f'A{current_row}'] = 'Baseline Comparisons'
                sheet[f'A{current_row}'].font = Font(bold=True)
                current_row += 2

        # Now write raw_data summary starting at the next available row
        if raw_data:
            pods_data = self._parse_json_response(raw_data) or {}
            if isinstance(pods_data, str):
                try:
                    pods_data = json.loads(pods_data)
                except Exception:
                    pods_data = {}
            elif not isinstance(pods_data, dict):
                pods_data = {}
            summary = pods_data.get("summary", {})
            
            # Summary information
            sheet[f'A{current_row}'] = "Total Pods:"
            sheet[f'B{current_row}'] = summary.get("total_pods", 0)
            current_row += 1
            
            sheet[f'A{current_row}'] = "Total Containers:"
            sheet[f'B{current_row}'] = summary.get("total_containers", 0)
            current_row += 1
            
            sheet[f'A{current_row}'] = "High CPU Pods:"
            sheet[f'B{current_row}'] = summary.get("high_cpu_pods", 0)
            current_row += 1
            
            sheet[f'A{current_row}'] = "High Memory Pods:"
            sheet[f'B{current_row}'] = summary.get("high_memory_pods", 0)
            current_row += 1

            # Per-Pod Usage Summary table
            individual_pods = (
                (pods_data.get('pods_total_usage') or {}).get('individual_pods')
                if isinstance(pods_data, dict) else {}
            )
 
            if isinstance(individual_pods, dict) and individual_pods:
                sheet[f'A{current_row}'] = 'Per-Pod Usage Summary'
                sheet[f'A{current_row}'].font = Font(bold=True)
                current_row += 1

                headers = [
                    'Namespace', 'Pod Name',
                    'CPU Min (%)', 'CPU Max (%)', 'CPU Mean (%)', 'CPU Samples',
                    'Memory Min (MB)', 'Memory Max (MB)', 'Memory Mean (MB)', 'Memory Samples'
                ]
                for col, header in enumerate(headers, 1):
                    sheet.cell(row=current_row, column=col, value=header)
                    sheet.cell(row=current_row, column=col).font = Font(bold=True)
                current_row += 1

                for pod_key, metrics in individual_pods.items():
                    namespace, pod_name = (pod_key.split('/', 1) + [""])[:2]
                    # Support both schemas: with nested 'statistics' or direct min/max/mean at top level
                    cpu_usage = (metrics.get('cpu_usage') or {}) if isinstance(metrics, dict) else {}
                    mem_usage = (metrics.get('memory_usage') or {}) if isinstance(metrics, dict) else {}
                    cpu_stats = cpu_usage.get('statistics') if isinstance(cpu_usage.get('statistics'), dict) else cpu_usage
                    mem_stats = mem_usage.get('statistics') if isinstance(mem_usage.get('statistics'), dict) else mem_usage
                    
                    values = [
                        namespace,
                        pod_name,
                        cpu_stats.get('min'),
                        cpu_stats.get('max'),
                        cpu_stats.get('mean'),
                        cpu_stats.get('count'),
                        mem_stats.get('min'),
                        mem_stats.get('max'),
                        mem_stats.get('mean'),
                        mem_stats.get('count'),
                    ]

                    for col, val in enumerate(values, 1):
                        sheet.cell(row=current_row, column=col, value=val)
                    current_row += 1

                # Adjust columns for readability
                for col_letter in ['A','B','C','D','E','F','G','H','I','J']:
                    sheet.column_dimensions[col_letter].width = 22

    def _create_network_sheet(self, sheet, network_analysis: Dict[str, Any], raw_data: Optional[Dict[str, Any]]):
        # Title
        sheet['A1'] = "Network Performance Analysis"
        sheet['A1'].font = Font(size=14, bold=True)

        current_row = 3
        if isinstance(network_analysis, dict) and network_analysis:
            health = network_analysis.get('health_analysis', {}) or {}
            bench = network_analysis.get('benchmark_comparison', {}) or {}
            summary = network_analysis.get('summary', {}) or {}

            # Analysis timestamp
            analysis_ts = network_analysis.get('analysis_timestamp', '')
            if analysis_ts:
                sheet[f'A{current_row}'] = 'Analysis Timestamp:'
                sheet[f'B{current_row}'] = analysis_ts
                current_row += 1

            # Key summary metrics
            health_score = health.get('overall_health_score', summary.get('health_score', ''))
            health_status = str(health.get('health_status', summary.get('overall_health', ''))).title()
            benchmark_rating = bench.get('overall_rating', summary.get('benchmark_rating', ''))

            if health_score != '':
                sheet[f'A{current_row}'] = 'Overall Health Score:'
                sheet[f'B{current_row}'] = health_score
                current_row += 1
            if health_status:
                sheet[f'A{current_row}'] = 'Health Status:'
                sheet[f'B{current_row}'] = health_status
                current_row += 1
            if benchmark_rating:
                sheet[f'A{current_row}'] = 'Benchmark Rating:'
                sheet[f'B{current_row}'] = benchmark_rating
                current_row += 2

            # Health analysis details
            if health:
                sheet[f'A{current_row}'] = 'Health Analysis'
                sheet[f'A{current_row}'].font = Font(bold=True)
                current_row += 1

                # Component scores table
                comp_scores = health.get('component_scores', {}) or {}
                if comp_scores:
                    sheet[f'A{current_row}'] = 'Component'
                    sheet[f'B{current_row}'] = 'Score'
                    sheet[f'A{current_row}'].font = Font(bold=True)
                    sheet[f'B{current_row}'].font = Font(bold=True)
                    current_row += 1
                    for comp, score in comp_scores.items():
                        sheet[f'A{current_row}'] = str(comp).replace('_', ' ').title()
                        sheet[f'B{current_row}'] = score
                        current_row += 1
                    current_row += 1

                # Issues list
                issues = health.get('issues', []) or []
                sheet[f'A{current_row}'] = 'Issues:'
                sheet[f'A{current_row}'].font = Font(bold=True)
                current_row += 1
                if issues:
                    for issue in issues[:20]:
                        sheet[f'A{current_row}'] = str(issue)
                        current_row += 1
                else:
                    sheet[f'A{current_row}'] = 'None'
                    current_row += 1
                current_row += 1

            # Capacity forecast
            capacity = network_analysis.get('capacity_forecast', {}) or {}
            if capacity:
                sheet[f'A{current_row}'] = 'Capacity Forecast'
                sheet[f'A{current_row}'].font = Font(bold=True)
                current_row += 1
                if 'forecast_period_days' in capacity:
                    sheet[f'A{current_row}'] = 'Forecast Period (days):'
                    sheet[f'B{current_row}'] = capacity.get('forecast_period_days')
                    current_row += 1
                if 'confidence' in capacity:
                    sheet[f'A{current_row}'] = 'Confidence:'
                    sheet[f'B{current_row}'] = capacity.get('confidence')
                    current_row += 2

            # Benchmark comparison
            if bench:
                sheet[f'A{current_row}'] = 'Benchmark Comparison'
                sheet[f'A{current_row}'].font = Font(bold=True)
                current_row += 1
                if 'overall_rating' in bench:
                    sheet[f'A{current_row}'] = 'Overall Rating:'
                    sheet[f'B{current_row}'] = bench.get('overall_rating')
                    current_row += 1
                if 'benchmark_date' in bench:
                    sheet[f'A{current_row}'] = 'Benchmark Date:'
                    sheet[f'B{current_row}'] = bench.get('benchmark_date')
                    current_row += 2

            # Summary block
            if summary:
                sheet[f'A{current_row}'] = 'Summary'
                sheet[f'A{current_row}'].font = Font(bold=True)
                current_row += 1
                if 'overall_health' in summary:
                    sheet[f'A{current_row}'] = 'Overall Health:'
                    sheet[f'B{current_row}'] = str(summary.get('overall_health', '')).title()
                    current_row += 1
                if 'health_score' in summary:
                    sheet[f'A{current_row}'] = 'Health Score:'
                    sheet[f'B{current_row}'] = summary.get('health_score')
                    current_row += 1
                if 'total_recommendations' in summary:
                    sheet[f'A{current_row}'] = 'Total Recommendations:'
                    sheet[f'B{current_row}'] = summary.get('total_recommendations')
                    current_row += 1
                if 'high_priority_recommendations' in summary:
                    sheet[f'A{current_row}'] = 'High Priority Recommendations:'
                    sheet[f'B{current_row}'] = summary.get('high_priority_recommendations')
                    current_row += 2

            # Key recommendations (brief)
            recs = network_analysis.get('optimization_recommendations', []) or []
            if recs:
                sheet[f'A{current_row}'] = 'Key Recommendations:'
                sheet[f'A{current_row}'].font = Font(bold=True)
                current_row += 1
                sheet[f'A{current_row}'] = 'Priority'
                sheet[f'B{current_row}'] = 'Recommendation'
                sheet[f'A{current_row}'].font = Font(bold=True)
                sheet[f'B{current_row}'].font = Font(bold=True)
                current_row += 1
                for rec in recs[:10]:
                    sheet[f'A{current_row}'] = rec.get('priority', '')
                    sheet[f'B{current_row}'] = rec.get('recommendation', '')
                    current_row += 1

        # Append raw data summary and alerts if available
        if raw_data:
            raw = raw_data
            if isinstance(raw_data, dict):
                raw = self._parse_json_response(raw_data) or raw_data
            # If still a JSON string, try to parse
            if isinstance(raw, str):
                try:
                    import json as _json
                    raw = _json.loads(raw)
                except Exception:
                    raw = {}

            if isinstance(raw, dict):
                has_summary = isinstance(raw.get('summary'), dict)
                has_alerts = isinstance(raw.get('alerts'), list)
                has_status = 'overall_status' in raw
                if has_summary or has_alerts or has_status:
                    # Section title
                    sheet[f'A{current_row}'] = 'Raw Network Summary'
                    sheet[f'A{current_row}'].font = Font(bold=True)
                    current_row += 1

                    if has_status:
                        sheet[f'A{current_row}'] = 'Overall Status:'
                        sheet[f'B{current_row}'] = raw.get('overall_status')
                        current_row += 1

                    summary_block = raw.get('summary') or {}
                    if isinstance(summary_block, dict) and summary_block:
                        mapping = [
                            ('Total Nodes', 'total_nodes'),
                            ('Total Interfaces', 'total_interfaces'),
                            ('Interfaces With Errors', 'interfaces_with_errors'),
                            ('Low Throughput Interfaces', 'low_throughput_interfaces'),
                            ('High Utilization Interfaces', 'high_utilization_interfaces'),
                        ]
                        for label, key in mapping:
                            sheet[f'A{current_row}'] = label
                            sheet[f'B{current_row}'] = summary_block.get(key)
                            current_row += 1
                        current_row += 1

                    if has_alerts:
                        alerts = raw.get('alerts') or []
                        if alerts:
                            # Alerts table
                            sheet[f'A{current_row}'] = 'Alerts'
                            sheet[f'A{current_row}'].font = Font(bold=True)
                            current_row += 1
                            headers = ['Type', 'Interface', 'Current (Mbps)', 'Baseline (Mbps)', 'Severity']
                            for col, header in enumerate(headers, 1):
                                sheet.cell(row=current_row, column=col, value=header)
                                sheet.cell(row=current_row, column=col).font = Font(bold=True)
                            for idx, alert in enumerate(alerts, current_row + 1):
                                sheet.cell(row=idx, column=1, value=alert.get('type'))
                                sheet.cell(row=idx, column=2, value=alert.get('interface'))
                                sheet.cell(row=idx, column=3, value=alert.get('current_mbps'))
                                sheet.cell(row=idx, column=4, value=alert.get('baseline_mbps'))
                                sheet.cell(row=idx, column=5, value=alert.get('severity'))
                            current_row = current_row + 1 + len(alerts)

        # Adjust column widths for readability
        for col in ['A', 'B', 'C', 'D', 'E']:
            sheet.column_dimensions[col].width = 35
            
    def _create_recommendations_sheet(self, sheet, analysis_results: Dict[str, Any]):
        """Create recommendations sheet."""
        sheet['A1'] = "Performance Recommendations"
        sheet['A1'].font = Font(size=14, bold=True)
        
        # Headers
        headers = ["Priority", "Component", "Issue", "Recommendation", "Impact"]
        for col, header in enumerate(headers, 1):
            sheet.cell(row=3, column=col, value=header)
            sheet.cell(row=3, column=col).font = Font(bold=True)
        
        # Collect all recommendations (robust to stringified JSON)
        row = 4
        seen = set()

        def write_rec(priority: str, component_name: str, issue: str, recommendation: str, impact: str):
            nonlocal row
            key = (priority or "", component_name or "", issue or "", recommendation or "", impact or "")
            if key in seen:
                return
            seen.add(key)
            sheet.cell(row=row, column=1, value=priority)
            sheet.cell(row=row, column=2, value=component_name)
            sheet.cell(row=row, column=3, value=issue)
            sheet.cell(row=row, column=4, value=recommendation)
            sheet.cell(row=row, column=5, value=impact)
            if priority == "high":
                for col_i in range(1, 6):
                    sheet.cell(row=row, column=col_i).fill = PatternFill(
                        start_color="FFCCCC", end_color="FFCCCC", fill_type="solid"
                    )
            elif priority == "medium":
                for col_i in range(1, 6):
                    sheet.cell(row=row, column=col_i).fill = PatternFill(
                        start_color="FFFFCC", end_color="FFFFCC", fill_type="solid"
                    )
            row += 1

        for component, analysis in analysis_results.items():
            parsed = analysis
            if isinstance(parsed, str):
                try:
                    parsed = json.loads(parsed)
                except Exception:
                    parsed = {}
            if not isinstance(parsed, dict):
                continue
            opt_recs = parsed.get("optimization_recommendations") or []
            if isinstance(opt_recs, list):
                for rec in opt_recs:
                    if not isinstance(rec, dict):
                        continue
                    write_rec(
                        priority=str(rec.get("priority", "")),
                        component_name=component,
                        issue=str(rec.get("issue", rec.get("category", ""))),
                        recommendation=str(rec.get("recommendation", "")),
                        impact=str(rec.get("impact", "")),
                    )
            # Also include plain-text recommendations from health_analysis
            health = parsed.get("health_analysis") or {}
            if isinstance(health, dict):
                for rec_text in (health.get("recommendations") or []):
                    write_rec(priority="", component_name=component, issue="", recommendation=str(rec_text), impact="")

        overall = analysis_results.get("overall_summary")
        if isinstance(overall, dict):
            for rec_text in (overall.get("recommendations") or []):
                write_rec(priority="", component_name="overall_summary", issue="", recommendation=str(rec_text), impact="")
        
        # Adjust column widths
        for col in range(1, 6):
            sheet.column_dimensions[chr(64 + col)].width = 25
    
    async def _generate_pdf_report(self, analysis_results: Dict[str, Any], filename: str) -> str:
        """Generate PDF report."""
        from fpdf.enums import XPos, YPos
        
        class PDF(FPDF):
            def header(self):
                self.set_font('Helvetica', 'B', 15)
                self.cell(0, 10, 'OpenShift Cluster Performance Report', new_x=XPos.LMARGIN, new_y=YPos.NEXT, align='C')
                self.ln(10)
        
        pdf = PDF()
        pdf.add_page()
        
        # Executive Summary
        pdf.set_font('Helvetica', 'B', 14)
        pdf.cell(0, 10, 'Executive Summary', new_x=XPos.LMARGIN, new_y=YPos.NEXT, align='L')
        pdf.ln(5)
        
        overall_summary = analysis_results.get("overall_summary", {}) or {}
        
        pdf.set_font('Helvetica', '', 12)
        pdf.cell(0, 8, f"Report Generated: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M:%S UTC')}", new_x=XPos.LMARGIN, new_y=YPos.NEXT)
        health_score = overall_summary.get('overall_health_score', 'Unknown')
        health_status = str(overall_summary.get('health_status', 'Unknown')).title()
        pdf.cell(0, 8, f"Overall Health Score: {health_score}", new_x=XPos.LMARGIN, new_y=YPos.NEXT)
        pdf.cell(0, 8, f"Health Status: {health_status}", new_x=XPos.LMARGIN, new_y=YPos.NEXT)
        pdf.ln(10)
        
        # Component Scores
        pdf.set_font('Helvetica', 'B', 12)
        pdf.cell(0, 8, 'Component Health Scores:', new_x=XPos.LMARGIN, new_y=YPos.NEXT)
        pdf.set_font('Helvetica', '', 10)
        
        component_scores = overall_summary.get("component_scores", {}) or {}
        if component_scores:
            for component, score in component_scores.items():
                pdf.cell(0, 6, f"  {component.replace('_', ' ').title()}: {score}", new_x=XPos.LMARGIN, new_y=YPos.NEXT)
        else:
            pdf.cell(0, 6, "  No component scores available", new_x=XPos.LMARGIN, new_y=YPos.NEXT)
        
        pdf.ln(10)
        
        # Critical Issues
        critical_issues = overall_summary.get("critical_issues", []) or []
        if critical_issues:
            pdf.set_font('Helvetica', 'B', 12)
            pdf.cell(0, 8, 'Critical Issues:', new_x=XPos.LMARGIN, new_y=YPos.NEXT)
            pdf.set_font('Helvetica', '', 10)
            
            for issue in critical_issues[:10]:  # Limit to first 10
                # Wrap long text
                wrapped_text = issue[:80] + "..." if len(issue) > 80 else issue
                # Use hyphen instead of bullet to avoid encoding issues
                pdf.cell(0, 6, f"  - {wrapped_text}", new_x=XPos.LMARGIN, new_y=YPos.NEXT)
        else:
            pdf.set_font('Helvetica', 'B', 12)
            pdf.cell(0, 8, 'Critical Issues:', new_x=XPos.LMARGIN, new_y=YPos.NEXT)
            pdf.set_font('Helvetica', '', 10)
            pdf.cell(0, 6, "  None detected", new_x=XPos.LMARGIN, new_y=YPos.NEXT)
        
        pdf.ln(10)
        
        # Recommendations
        pdf.set_font('Helvetica', 'B', 12)
        pdf.cell(0, 8, 'Key Recommendations:', new_x=XPos.LMARGIN, new_y=YPos.NEXT)
        pdf.set_font('Helvetica', '', 10)
        
        recommendations = overall_summary.get("recommendations", []) or []
        if recommendations:
            for rec in recommendations[:15]:  # Limit to first 15
                text = str(rec)
                wrapped_rec = text[:80] + "..." if len(text) > 80 else text
                pdf.cell(0, 6, f"  - {wrapped_rec}", new_x=XPos.LMARGIN, new_y=YPos.NEXT)
        else:
            pdf.cell(0, 6, "  No recommendations available", new_x=XPos.LMARGIN, new_y=YPos.NEXT)
        
        # Save PDF
        Path("exports").mkdir(exist_ok=True)
        pdf.output(filename)
        logger.info(f"PDF report saved: {filename}")
        
        return filename
    
    async def run_analysis(self, task: str = "comprehensive_analysis") -> Dict[str, Any]:
        """Run the complete analysis workflow."""
        initial_state = {
            "messages": [],
            "current_task": task,
            "collected_data": {},
            "analysis_results": {},
            "report_generated": False,
            "error": None
        }
        
        # Run the graph
        config = {"configurable": {"thread_id": "analysis_session"}}
        final_state = await self.graph.ainvoke(initial_state, config=config)
        
        return final_state


async def main():
    """Main entry point for the agent."""
    # Check for required environment variables
    openai_api_key = os.getenv("OPENAI_API_KEY")
    if not openai_api_key:
        logger.error("OPENAI_API_KEY environment variable is required")
        sys.exit(1)
    
    kubeconfig = os.getenv("KUBECONFIG")
    if not kubeconfig:
        logger.error("KUBECONFIG environment variable is required")
        sys.exit(1)
    
    # Create agent
    agent = OCPBenchmarkAgent(openai_api_key)
    
    try:
        logger.info("Starting OpenShift Benchmark AI Agent...")
        
        # Run comprehensive analysis
        result = await agent.run_analysis("comprehensive_analysis")
        
        if result.get("error"):
            logger.error(f"Analysis failed: {result['error']}")
            return 1
        
        if result.get("report_generated"):
            logger.info("Analysis completed successfully!")
            logger.info("Check the exports/ directory for generated reports")
            
            # Print summary
            overall_summary = result.get("analysis_results", {}).get("overall_summary", {})
            print("\n" + "="*60)
            print("OPENSHIFT CLUSTER PERFORMANCE SUMMARY")
            print("="*60)
            print(f"Overall Health Score: {overall_summary.get('overall_health_score', 'Unknown')}")
            print(f"Health Status: {overall_summary.get('health_status', 'Unknown').title()}")
            print(f"Critical Issues: {len(overall_summary.get('critical_issues', []))}")
            print(f"Total Recommendations: {len(overall_summary.get('recommendations', []))}")
            print("="*60)
            
            return 0
        else:
            logger.error("Analysis completed but no reports were generated")
            return 1
    
    except Exception as e:
        logger.error(f"Agent execution failed: {e}")
        return 1


if __name__ == "__main__":
    try:
        exit_code = asyncio.run(main())
        sys.exit(exit_code)
    except KeyboardInterrupt:
        logger.info("Agent execution interrupted")
        sys.exit(130)
    except Exception as e:
        logger.error(f"Unexpected error: {e}")
        sys.exit(1)